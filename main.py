import os
import sys
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QFileDialog, \
    QLabel, QSpinBox, QMessageBox

import excel_handler  # Убедитесь, что ваш файл excel_handler.py доступен


class ExcelMergerApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Обработка Excel файлов")
        self.setGeometry(300, 100, 500, 250)
        self.setFixedSize(500, 250)  # Фиксированный размер окна
        self.init_ui()

    def init_ui(self):
        # Основной вертикальный контейнер
        main_layout = QVBoxLayout()

        # Панель для ввода папки источника
        source_layout = QHBoxLayout()
        self.source_label = QLabel("Источник файлов:")
        self.source_input = QLineEdit()
        self.source_button = QPushButton("Выбрать папку")
        self.source_button.clicked.connect(self.browse_source_directory)
        source_layout.addWidget(self.source_label)
        source_layout.addWidget(self.source_input)
        source_layout.addWidget(self.source_button)

        # Панель для ввода URL
        url_layout = QHBoxLayout()
        self.url_label = QLabel("URL для получения данных:")
        self.url_input = QLineEdit()
        url_layout.addWidget(self.url_label)
        url_layout.addWidget(self.url_input)

        # Панель для ввода лимита
        limit_layout = QHBoxLayout()
        self.limit_label = QLabel("Лимит (число):")
        self.limit_input = QSpinBox()
        self.limit_input.setRange(1, 1000)  # Устанавливаем диапазон для лимита
        limit_layout.addWidget(self.limit_label)
        limit_layout.addWidget(self.limit_input)

        # Кнопка выбора места сохранения файла
        dest_layout = QHBoxLayout()
        self.dest_button = QPushButton("Выбрать файл для сохранения")
        self.dest_button.clicked.connect(self.browse_destination_file)
        dest_layout.addWidget(self.dest_button)

        # Кнопка запуска
        self.start_button = QPushButton("Запустить")
        self.start_button.clicked.connect(self.on_start_button_click)

        # Добавление всех панелей на главный layout
        main_layout.addLayout(source_layout)
        main_layout.addLayout(url_layout)
        main_layout.addLayout(limit_layout)
        main_layout.addLayout(dest_layout)
        main_layout.addWidget(self.start_button)

        self.setLayout(main_layout)

        # Применение темной темы
        self.setStyleSheet("""
            QWidget {
                background-color: #2E2E2E;
                color: white;
            }
            QLabel {
                color: white;
            }
            QLineEdit, QSpinBox {
                background-color: #444444;
                color: white;
                border: 1px solid #666666;
            }
            QPushButton {
                background-color: #444444;
                color: white;
                border: 1px solid #666666;
                padding: 5px 10px;
            }
            QPushButton:hover {
                background-color: #555555;
            }
            QSpinBox {
                background-color: #444444;
                color: white;
                border: 1px solid #666666;
                width: 80px;
            }
        """)

    def browse_source_directory(self):
        folder = QFileDialog.getExistingDirectory(self, "Выбрать папку источника")
        if folder:
            self.source_input.setText(folder)

    def browse_destination_file(self):
        # Открываем диалог выбора имени и места сохранения файла
        file_path, _ = QFileDialog.getSaveFileName(self, "Выбрать файл для сохранения", "", "Excel Files (*.xlsx)")
        if file_path:
            self.destination_file = file_path  # Сохраняем путь к выбранному файлу

    def start_processing(self, source_directory, destination_file, url, limit):
        try:
            # Запуск excel_handler.main_excel с введенными параметрами
            excel_handler.main_excel(source_directory, limit + 1, url)
            self.copy_data_to_single_file(source_directory, destination_file)
            QMessageBox.information(self, "Успех", "Процесс завершен успешно!")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {e}")

    def copy_data_to_single_file(self, source_directory, destination_file):
        dest_wb = openpyxl.Workbook()
        dest_ws = dest_wb.active
        current_row = 2
        excel_handler.add_header(dest_ws)
        for i in range(1, 150):
            filename = f'invest{i}.xlsx'
            file_path = os.path.join(source_directory, filename)
            if os.path.exists(file_path):
                source_wb = openpyxl.load_workbook(file_path)
                source_ws = source_wb.active
                for row in source_ws.iter_rows(min_row=2, max_row=source_ws.max_row):
                    for cell in row:
                        dest_ws.cell(row=current_row, column=cell.col_idx, value=cell.value)
                    current_row += 1
                source_wb.close()
        dest_wb.save(f"{destination_file}.xlsx")
        dest_wb.close()

    def on_start_button_click(self):
        source_directory = self.source_input.text()
        destination_file = getattr(self, 'destination_file', None)
        url = self.url_input.text()
        limit = self.limit_input.value()
        if not source_directory or not destination_file or not url:
            QMessageBox.warning(self, "Предупреждение", "Пожалуйста, заполните все поля.")
            return
        self.start_processing(source_directory, destination_file, url, limit)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelMergerApp()
    window.show()
    sys.exit(app.exec_())
